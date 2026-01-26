


from bs4 import BeautifulSoup
import urllib.request
import requests
import pandas as pd
import openpyxl
from datetime import datetime
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders






file_name = "D:\Popmart-ThousandIsland.xlsx"  # Replace with your Excel file name
input_sheet_name = "Series"  # Replace with the sheet name where the links are stored
output_sheet_name = "tracking"

# Read the Excel file
df = pd.read_excel(file_name, sheet_name=input_sheet_name)


headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.110 Safari/537.36"
}

# Create an email message body
email_body = "Here is the extraction status of each URL:\n\n"


#打开每个系列的网站
for index, row in df.iterrows():
    url = row['URL']
    series = row['Series']

    response = requests.get(url, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')

    #读取每个系列里面的SKU
    divs = soup.find_all('div', class_='mb-8px')
    a_tags = soup.find_all('a', {'data-v-ba00a76c': "", 'aria-current': "page"})
    # Prepare data for Excel
    data_to_append = []
    current_date = datetime.now().strftime("%Y-%m-%d")  # Get current date
    base_url = "https://qiandao.com"  # Base URL to append

    
    for tag in a_tags:
        href = tag.get('href')
        full_url = base_url + href  # Concatenate the base URL with the href
        h3_tag = tag.find('h3', {'data-v-ba00a76c': "", 'class': 'name u-otext2'})
        h3_text = h3_tag.get_text(strip=True) if h3_tag else "N/A"
        p_tag = tag.find('p', {'data-v-ba00a76c': "", 'class': 'wish-desc'})
        p_text = p_tag.get_text(strip=True) if p_tag else "N/A"
        div_tag = tag.find('div', {'data-v-ba00a76c': "", 'class': 'spu-price text-center font-num text-n6'})
        div_text = div_tag.get_text(strip=True) if div_tag else "N/A"
        data_to_append.append([series, h3_text, p_text, div_text, full_url, current_date])

    # Append to Excel
    try:
        # Load the workbook if it exists, otherwise create a new one
        try:
            workbook = openpyxl.load_workbook(file_name)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # Get or create the sheet
        if output_sheet_name in workbook.sheetnames:
            sheet = workbook[output_sheet_name]
        else:
            sheet = workbook.create_sheet(output_sheet_name)

        # Add headers if the sheet is empty
        if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
            headers = ["系列名","商品名", "想要人数", "交易价格", "网址", "日期","正在出售","正在求购","付款人数","成交均价"]
            sheet.append(headers)

        # Append data
        for row in data_to_append:
            sheet.append(row)

        # Save the workbook
        workbook.save(file_name)
        print(f"Data appended successfully to {file_name} in sheet '{output_sheet_name}'.")

        # Append the status to the email body
        email_body += f"URL: {url}\nSeries: {series}\nStatus: Success\n\n"

    except Exception as e:
        print(f"An error occurred: {e}")
        # Append the status to the email body
        email_body += f"URL: {url}\nSeries: {series}\nStatus: Failed\nError: {str(e)}\n\n"
    
    time.sleep(1)

        

    
def send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path):
    # Compose the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Attach the message body
    msg.attach(MIMEText(message, 'plain'))

    # Attach the file
    attachment = open(attachment_path, 'rb')
    filename = attachment_path.split('/')[-1]  # Extracts the filename from the path
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {filename}')
    msg.attach(part)

    # Connect to the email server and send the email
    with smtplib.SMTP('smtp.qq.com', 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

    print('Email sent successfully!')



sender_email = '396481139@qq.com'
sender_password = 'mocjzkhznmudbghf'
receiver_email = 'cuiyuan@maisoncapital.com,linhuaqiang@maisoncapital.com,zengleshi@maisoncapital.com,cuiwenli@maisoncapital.com,396481139@qq.com'

subject = '泡泡玛特二手数据-系列'+ current_date


message = email_body
print(message)
attachment_path = file_name
send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)






   






# coding=gbk

from bs4 import BeautifulSoup
import urllib.request
import requests
from urllib import request
import time, datetime, subprocess, random
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import datetime
import openpyxl

#delay = random.uniform(0, 600)
#time.sleep(delay)


def sleep_random_second(num):
    while num < 5:
    	# random.uniform����ʵ�ֳ�˯0.01-0.001�룬��Ҫround����֤��Ч����
        time.sleep(round(random.uniform(1, 0.5), 3))
        print("========================", round(random.uniform(1,0.5), 3))
        num = num + 1

url='https://www.58moto.com/ranks/car/0/index.html'

headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

response = requests.get(url, headers=headers)
print(response.status_code)

soup = BeautifulSoup(response.text, 'html.parser')
print(soup.prettify())
table_element = soup.find('div', class_='content')
rows = table_element.find_all('a', target="_blank")
file_path = r"D:\58moto.xlsx"  # Use raw string to handle backslashes in file path


##�½�excel�ļ�
#workbook = openpyxl.Workbook()
#sheet = workbook.active
#sheet.title = "Extracted Data"
#sheet.append(["ranking", "��������", "Box Ellipsis Text", "Box Text 1", "Box Text 2", "Box Text 3","Date"]) 

current_date = datetime.datetime.now().strftime("%Y-%m-%d")  # Format: YYYY-MM-DD

#������excel�ļ�
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Prepare the plain text content for the email
email_body = "Scraped Data from 58Moto\n\n"
email_body += "Ranking | ���� | ���� | ����ֵ | Date | ����\n"
email_body += "-" * 80 + "\n"  # Add a separator line



for rank, row in enumerate(rows, start=1):

    # Extract text from the row itself
    row_text = row.get_text(strip=True)

    # Extract data from divs with class "box ellipsis"
    box_ellipsis = row.find('div', class_='box ellipsis')
    box_ellipsis_text = box_ellipsis.get_text(strip=True) if box_ellipsis else ""

    # Extract data from all divs with class "box"
    box_elements = row.find_all('div', class_='box')
    box_texts = [box.get_text(strip=True) for box in box_elements]

    # Ensure the number of boxes match the expected columns in the Excel file
    # Fill missing boxes with empty strings if there are fewer box elements
    while len(box_texts) < 3:  # Adjust 3 to the maximum number of box columns you're expecting
        box_texts.append("")
    # Append the extracted data to the sheet
    sheet.append([rank] + box_texts[:3]+ [current_date] + [row_text])  # Limiting to 3 box columns
    # Add the row to the email body
    email_body += f"{rank} | {box_texts[0]} | {box_texts[1]} | {box_texts[2]} | {current_date} | {row_text}\n"


# Save the Excel file to the specified file path
workbook.save(file_path)
print(f"Data saved to '{file_path}'")


def send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path):
    # Compose the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Attach the message body
    msg.attach(MIMEText(message, 'plain'))

    # Attach the file
    attachment = open(attachment_path, 'rb')
    filename = attachment_path.split('/')[-1]  # Extracts the filename from the path
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {filename}')
    msg.attach(part)

    # Connect to the email server and send the email
    with smtplib.SMTP('smtp.qq.com', 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

    print('Email sent successfully!')



email_date = datetime.date.today()
formatted_date = email_date.strftime("%Y-%m-%d")

sender_email = '396481139@qq.com'
sender_password = 'mocjzkhznmudbghf'
receiver_email = 'cuiyuan@maisoncapital.com,linhuaqiang@maisoncapital.com'

subject = 'Ħ�з���������'+ formatted_date




message = '��������Ϊ: ' + email_body
print(message)

attachment_path = file_path

# send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)

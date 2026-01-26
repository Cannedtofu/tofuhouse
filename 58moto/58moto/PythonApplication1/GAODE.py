# coding=gbk

from bs4 import BeautifulSoup
import urllib.request
from urllib import request
import time, datetime, subprocess, random
from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import datetime

#delay = random.uniform(0, 600)
#time.sleep(delay)


def sleep_random_second(num):
    while num < 5:
    	# random.uniform可以实现沉睡0.01-0.001秒，需要round来保证有效数字
        time.sleep(round(random.uniform(1, 0.5), 3))
        print("========================", round(random.uniform(1,0.5), 3))
        num = num + 1

url='https://report.amap.com/diagnosis/index.do'

driver = webdriver.Chrome()
driver.get(url)
element = driver.find_element(By.CLASS_NAME, "main") #找到使用必读的部分
sleep_random_second(3)
driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", element)#使用必读部分拉到最下面
sleep_random_second(4)
checkboxes = driver.find_elements(By.XPATH,"//input[@type='checkbox']")#找到三个条款同意
for checkbox in checkboxes: #每个条款打钩
    checkbox.click()
    sleep_random_second(3)
button = driver.find_element(By.ID,"SURE_BUTTON") #点击提交
button.click()

sleep_random_second(1)
sleep_random_second(1)
page_source = driver.page_source #导出网页中的html,作为string输出


soup = BeautifulSoup(page_source, 'html.parser')
table_element = soup.find('table', class_='table')


data = [] #输出表格至数据结构
if table_element:
    rows = table_element.find_all('tr')
    for row in rows:
        row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
        data.append(row_data)

print(data)

file_path = 'D:\GaoDe.xlsx' #定义目标文件
df = pd.DataFrame(data)
df_final = df.transpose() #表格调整transpose
df_final = df_final.drop(0) #去除首行
df_final = df_final.values.reshape(1, -1) #变成单行
df_final = pd.DataFrame(df_final)

current_time = pd.Timestamp.now() #加入时间戳
data = [[str(current_time)] + df_final.values.flatten().tolist()]
df_final = pd.DataFrame(data, columns=['Time'] + df_final.columns.tolist())

book = load_workbook(file_path)
sheet = book.active
start_row = sheet.max_row + 1 #找到数据最后一行

for row in df_final.iterrows():
    sheet.append(row[1].tolist())

book.save(file_path)
book.close()




def send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path):
    # Compose the email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = subject

    # Attach the message body
    msg.attach(MIMEText(message, 'plain'))

    # Attach the file
    attachment = open(attachment_path, 'rb')
    filename = attachment_path.split('/')[-1]  # Extracts the filename from the path
    part = MIMEBase('application', 'octet-stream')
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {filename}')
    msg.attach(part)

    # Connect to the email server and send the email
    with smtplib.SMTP('smtp.qq.com', 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

    print('Email sent successfully!')


current_date = datetime.date.today()
formatted_date = current_date.strftime("%Y-%m-%d")

sender_email = '396481139@qq.com'
sender_password = 'mocjzkhznmudbghf'
receiver_email = 'cuiyuan@maisoncapital.com,zhj@maisoncapital.com'


subject = '高德地图数据'+ formatted_date
if float(data[0][3].rstrip('%'))>0:
    message = '今日数据为: ' + str(data[0])
else:
    message = '今日数据爬取失败'
attachment_path = file_path

send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)

# -*- coding: utf-8 -*-

from bs4 import BeautifulSoup
import urllib.request
from urllib import request
import time, datetime, subprocess, random
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import datetime
import os

#delay = random.uniform(0, 600)
#time.sleep(delay)

options = Options()
options.add_argument("--ignore-certificate-errors")
options.add_argument("--start-maximized")
options.add_argument("--ignore-ssl-errors")
options.add_argument("--disable-web-security")
options.add_argument("--headless")  # 开启无头模式
options.add_argument("--no-sandbox")  # 规避 Linux 权限问题
options.add_argument("--disable-dev-shm-usage")  # 解决内存限制问题
options.add_argument("--window-size=1920,1080")  # 固定分辨率防止元素找不到
options.add_experimental_option("excludeSwitches", ["enable-logging"])


# Suppress Chrome's own noisy logs
options.add_experimental_option("excludeSwitches", ["enable-logging"])
options.add_argument("--log-level=3")  # 0=ALL, 1=INFO, 2=WARNING, 3=ERROR

service = Service(ChromeDriverManager().install())
service.log_path = "NUL"  # discard chromedriver logs (use "/dev/null" on Linux/macOS)

driver = webdriver.Chrome(service=service, options=options)





def sleep_random_second(num):
    while num < 5:
    	# random.uniform����ʵ�ֳ�˯0.01-0.001�룬��Ҫround����֤��Ч����
        time.sleep(round(random.uniform(1, 0.5), 3))
        print("========================", round(random.uniform(1,0.5), 3))
        num = num + 1

url='https://report.amap.com/diagnosis/index.do'


driver.get(url)
element = driver.find_element(By.CLASS_NAME, "main") #�ҵ�ʹ�ñض��Ĳ���
sleep_random_second(3)
driver.execute_script("arguments[0].scrollTop = arguments[0].scrollHeight;", element)#ʹ�ñض���������������
sleep_random_second(4)
checkboxes = driver.find_elements(By.XPATH,"//input[@type='checkbox']")#�ҵ���������ͬ��
for checkbox in checkboxes: #ÿ�������
    checkbox.click()
    sleep_random_second(3)
button = driver.find_element(By.ID,"SURE_BUTTON") #����ύ
button.click()

sleep_random_second(1)
sleep_random_second(1)
page_source = driver.page_source #������ҳ�е�html,��Ϊstring���


soup = BeautifulSoup(page_source, 'html.parser')
table_element = soup.find('table', class_='table')


data = [] #������������ݽṹ
if table_element:
    rows = table_element.find_all('tr')
    for row in rows:
        row_data = [cell.get_text(strip=True) for cell in row.find_all('td')]
        data.append(row_data)

print(data)

file_path = './GaoDe.xlsx' #����Ŀ���ļ�
df = pd.DataFrame(data)
df_final = df.transpose() #�������transpose
df_final = df_final.drop(0) #ȥ������
df_final = df_final.values.reshape(1, -1) #��ɵ���
df_final = pd.DataFrame(df_final)

current_time = pd.Timestamp.now() #����ʱ���
data = [[str(current_time)] + df_final.values.flatten().tolist()]
df_final = pd.DataFrame(data, columns=['Time'] + df_final.columns.tolist())

book = load_workbook(file_path)
sheet = book.active
start_row = sheet.max_row + 1 #�ҵ��������һ��

for row in df_final.iterrows():
    sheet.append(row[1].tolist())

book.save(file_path)
book.close()


def send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path):
    # 1. 处理收件人：将字符串转换为列表，用于 SMTP 发送
    receivers = [addr.strip() for addr in receiver_email.split(',')]

    # 2. 构建邮件内容
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email  # Header 中保持逗号分隔的字符串
    msg['Subject'] = subject

    msg.attach(MIMEText(message, 'plain'))

    # 3. 添加附件
    try:
        with open(attachment_path, 'rb') as attachment:
            filename = os.path.basename(attachment_path)
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename={filename}')
            msg.attach(part)
    except Exception as e:
        print(f"附件读取失败: {e}")
        return

    # 4. 连接服务器并发送 (改用 465 端口和 SMTP_SSL)
    try:
        # 使用 SMTP_SSL 建立安全连接
        with smtplib.SMTP_SSL('smtp.qq.com', 465) as smtp:
            # smtp.set_debuglevel(1)  # 如果仍然失败，取消此行注释查看详细日志
            smtp.login(sender_email, sender_password)
            # 指定 to_addrs 为收件人列表
            smtp.send_message(msg, from_addr=sender_email, to_addrs=receivers)
        print('Email sent successfully!')
    except smtplib.SMTPException as e:
        print(f'Email sending failed: {e}')






current_date = datetime.date.today()
formatted_date = current_date.strftime("%Y-%m-%d")

sender_email = '396481139@qq.com'
sender_password = 'mocjzkhznmudbghf'
receiver_email = 'cuiyuan@maisoncapital.com,zhj@maisoncapital.com,396481139@qq.com,huaqianglin88@126.com'


subject = '高德地图数据'+ formatted_date
if float(data[0][3].rstrip('%'))>0:
    message = '今日数据为: ' + str(data[0])
else:
    message = '获取数据失败'
attachment_path = file_path

send_email(sender_email, sender_password, receiver_email, subject, message, attachment_path)

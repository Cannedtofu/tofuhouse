from bs4 import BeautifulSoup
import urllib.request
import requests
from selenium import webdriver
import openpyxl
import time
import os
from openpyxl import Workbook

excel_location = "D:\\Arcteryx brand store_202512.xlsx"

if os.path.exists(excel_location):
    # Open existing file
    wb = openpyxl.load_workbook(excel_location)
    # Get sheet named "Data", create if missing
    if "Data" in wb.sheetnames:
        sheet = wb["Data"]
    else:
        sheet = wb.create_sheet("Data")
else:
    # Create new workbook and sheet
    wb = Workbook()
    # Remove the default "Sheet"
    default_sheet = wb.active
    wb.remove(default_sheet)
    # Add "Data" sheet
    sheet = wb.create_sheet("Data")
    # Save the new file
    wb.save(excel_location)


file_path = r'D:\远程云盘\SynologyDrive\麦星(远程)\研究\亚玛芬\爬虫\202512\official store_202512.txt'
with open(file_path, 'r', encoding='utf-8') as file:
    content = file.read()

soup = BeautifulSoup(content, "html.parser")


data = soup.find_all("a")

for store in data:
    output = []
    store_name = store.find('div',class_='index-location-name').text
    store_city = store.find('div',class_='index-location-desc').text
    store_type = store.get('data-categories')
    store_url = 'https://stores.arcteryx.com/' + store.get('href')
    store_address = ''
    
    # response = requests.get(store_url)
    # html_content = response.text
    # soup = BeautifulSoup(html_content, 'html.parser')
    # try:
    #     store_address = soup.find('span',class_='landing-header-detail-item').text
    # except:
    #     store_address = ''

    print([store_name,store_city,store_address,store_type,store_url])
    sheet.append([store_name,store_city,store_address,store_type,store_url])
    wb.save(excel_location)
    # time.sleep(0.2)  # Sleep to avoid overwhelming the server
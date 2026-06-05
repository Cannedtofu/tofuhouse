from bs4 import BeautifulSoup
import urllib.request
import requests
import pandas as pd
import openpyxl
from datetime import datetime
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def setup_driver():
    """Initialize and return a Chrome WebDriver with optimized settings"""
    chrome_options = Options()
    
    # Essential options for stability
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--start-maximized")
    
    # Anti-detection measures
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_experimental_option('useAutomationExtension', False)
    
    # SSL and certificate handling
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--ignore-ssl-errors")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-web-security")
    
    # Performance optimizations
    chrome_options.add_argument("--disable-extensions")
    chrome_options.add_argument("--disable-plugins")
    chrome_options.add_argument("--disable-images")  # Optional: disable images for faster loading
    
    # User agent
    chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
    
    try:
        # Try using webdriver-manager first (recommended)
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        logger.info("Driver initialized with webdriver-manager")
    except Exception as e:
        logger.warning(f"webdriver-manager failed: {e}")
        try:
            # Fallback to your local driver path
            service = Service(r'C:\Program Files\python chrome\chrome-win\chromedriver.exe')
            driver = webdriver.Chrome(service=service, options=chrome_options)
            logger.info("Driver initialized with local chromedriver")
        except Exception as e:
            logger.error(f"Failed to initialize driver: {e}")
            raise
    
    # Configure timeouts
    driver.set_page_load_timeout(30)
    driver.implicitly_wait(10)
    
    # Execute script to hide automation indicators
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    
    return driver

def safe_get_page(driver, url, max_retries=1):
    """Safely get a page with retry logic"""
    for attempt in range(max_retries):
        try:
            logger.info(f"Loading URL (attempt {attempt + 1}): {url}")
            driver.get(url)
            
            # Wait for page to load
            WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            return driver.page_source
            
        except TimeoutException:
            logger.warning(f"Timeout loading {url} on attempt {attempt + 1}")
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            else:
                logger.error(f"Failed to load {url} after {max_retries} attempts")
                return driver.page_source
                
        except WebDriverException as e:
            logger.error(f"WebDriver error on attempt {attempt + 1}: {e}")
            if attempt < max_retries - 1:
                time.sleep(2)
                continue
            else:
                raise


# Initialize driver globally
driver = setup_driver()

def scrap_from_list(url):
    data=[]

    try:
        safe_get_page(driver,url)
        page_source = driver.page_source

    except TimeoutException:
        page_source = driver.page_source

    soup = BeautifulSoup(page_source, 'html.parser')
    dealer_item = soup.find_all('div',{'class':"dealer-item"})
    
    for dealer in dealer_item:
        dealer_url=dealer.find('a')['href']
        dealer_url = 'https://www.polaris.com'+ dealer_url
        dealer_city_tag = dealer.find('a',{'class':"dealer-city font-family-primary font-size-sm margin-bottom-xs display-block" })
        dealer_city = dealer_city_tag.get_text(strip=True)if dealer_city_tag else "N/A"
        dealer_name_tag=dealer.find('div',{'class':'dealer-name font-color-primary text-decoration-none' })
        dealer_name = dealer_name_tag.get_text(strip=True)if dealer_name_tag else "N/A"
        dealer_address_tag = dealer.find('address',{'class':'dealer-address font-family-default font-size-sm' })
        dealer_address= dealer_address_tag.get_text(strip=True)if dealer_address_tag else "N/A"
        dealer_phone_tag = dealer.find('div',{'class':'pull-left margin-right-xs'})
        dealer_phone= dealer_phone_tag.get_text(strip=True)if dealer_phone_tag else "N/A"
        


        data.append([dealer_url,dealer_name,dealer_address,dealer_phone,dealer_city])
        print([dealer_url,dealer_name,dealer_address,dealer_phone,dealer_city])
         

    
    print(data)
    time.sleep(1)

    return data

def get_dealer_url(city_url):

    try:
        safe_get_page(driver,city_url)
        page_source = driver.page_source
    except TimeoutException:
        page_source = driver.page_source

    soup = BeautifulSoup(page_source, 'html.parser')
    items = soup.find_all('a',{'class':"state link-decoration-type-simple font-family-default font-size-sm font-color-primary display-block"})

    data=[]

    for state in items:
        state_url = state.get('href')
        state_name = state.text
        state_url = 'https://www.polaris.com'+ state_url
    
        print(state_name,"  ",state_url)

        try:
            safe_get_page(driver,state_url)
            page_source = driver.page_source
        except TimeoutException:
            page_source = driver.page_source

        #dealers
        soup = BeautifulSoup(page_source, 'html.parser')
        dealer_item = soup.find_all('div',{'class':"dealer-item"})
    
        for dealer in dealer_item:
            dealer_url=dealer.find('a')['href']
            dealer_city_tag = dealer.find('a',{'class':"dealer-city font-family-primary font-size-sm margin-bottom-xs display-block" })
            dealer_city = dealer_city_tag.get_text(strip=True)if dealer_city_tag else "N/A"
            dealer_name_tag=dealer.find('div',{'class':'dealer-name font-color-primary text-decoration-none' })
            dealer_name = dealer_name_tag.get_text(strip=True)if dealer_name_tag else "N/A"
            dealer_address_tag = dealer.find('address',{'class':'dealer-address font-family-default font-size-sm' })
            dealer_address= dealer_address_tag.get_text(strip=True)if dealer_address_tag else "N/A"
            dealer_phone_tag = dealer.find('div',{'class':'pull-left margin-right-xs'})
            dealer_phone= dealer_phone_tag.get_text(strip=True)if dealer_phone_tag else "N/A"
        

            #safe_get_page(driver,dealer_url)
            #page_source = driver.page_source
            #soup = BeautifulSoup(page_source, 'html.parser')
            #try:
            #    dealer_detail_address = soup.find('address').text
            #except:
            #    print('fail to get brands')
            #try:
            #    dealer_brands= soup.find('div', class_='multiple-brands-support margin-top-sm').text
            #except:
            #    print('fail to get brands')

            data.append([dealer_url,dealer_name,dealer_address,dealer_phone,dealer_city])
            print([dealer_url,dealer_name,dealer_address,dealer_phone,dealer_city])
        

    
        print(data)
        time.sleep(1)

    return data

def scrap_city(dealer_url):
    data=[]
    dealer_url,dealer_detail_address,dealer_brands="","",""
    soup=None
    
    try:
        print("start loading")
        safe_get_page(driver,dealer_url)
        print("finish loading")
        page_source = driver.page_source

    except TimeoutException:
        page_source = driver.page_source
    except Exception as e:
        print("Error loading page:", e)
        soup=None
    
    
    soup = BeautifulSoup(page_source, 'html.parser')

    try:
        dealer_detail_address = soup.find('address').text
    except:
        print('fail to get brands')
        dealer_detail_address=""
    try:
        dealer_brands= soup.find('div', class_='multiple-brands-support margin-top-sm').text
    except:
        print('fail to get brands')
        dealer_brands=""

    data=[dealer_url,dealer_detail_address,dealer_brands]
    print(dealer_url,dealer_detail_address,dealer_brands)
    time.sleep(0.3)

    
    print(data)



    return data

def append_to_excel(file_name,sheet_name,data_to_append):
    #try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook[sheet_name]
        if any(isinstance(item, list) for item in data_to_append):
            for row in data_to_append:
                sheet.append(row)
        else:
            sheet.append(data_to_append)

        workbook.save(file_name)
        print(f"Data appended successfully to {file_name} in sheet '{sheet_name}'.")

    #except Exception as e:
    #    print("failed to input data")

def clean_info(info_list):
    return [item.replace("  ", "").replace("\n", "") for item in info_list]


file_name = "D:\\Polaris Dealer.xlsx"  # Replace with your Excel file name

# #--------Part 1-----------
# #get state general list

# # Set the page load timeout to 3 second
# # driver.set_page_load_timeout(10)

# output_sheet_name = "Url"
# data = get_dealer_url('https://www.polaris.com/en-us/off-road/dealers/')
# append_to_excel(file_name,output_sheet_name,data)



#---------------Part 2------------------
# set the page load timeout to x second
driver.set_page_load_timeout(5)

input_sheet_name = "Url"  # Replace with the sheet name where the links are stored
output_sheet_name = "Dealer"
workbook = openpyxl.load_workbook(file_name)
sheet = workbook[input_sheet_name]

for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1), start=1):
    dealer_url=row[0].value
    print(dealer_url)
    type(dealer_url)
    data=[]
    data = scrap_city(dealer_url)
    data = clean_info(data)
    data_to_append =[dealer_url, data[1], data[2]]
    print("current progress: ",index,"/",sheet.max_row)
    append_to_excel("D:\\Polaris Dealer.xlsx",output_sheet_name,data_to_append)



# #---------------Part 3------------------ 有的是二级清单，part2爬的时候爬不到，需要再跑一次part3
# # set the page load timeout to x second
# driver.set_page_load_timeout(5)


# input_sheet_name = "Url"  # Replace with the sheet name where the links are stored
# output_sheet_name = "Dealer"
# workbook = openpyxl.load_workbook(file_name)
# sheet = workbook[input_sheet_name]

# for index, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1), start=1):
#    list_page_url=row[0].value

#    print(list_page_url)
#    dealer_list=scrap_from_list(list_page_url)
#    print(dealer_list)

#    for dealer in dealer_list:
#        dealer_url=dealer[0]
#        print("scraping : *** "+dealer_url)

#        data=[]
#        data = scrap_city(dealer_url)
#        data = clean_info(data)
#        data_to_append =[dealer_url, data[1], data[2]]

#        append_to_excel("D:\\Polaris Dealer.xlsx",output_sheet_name,data_to_append)
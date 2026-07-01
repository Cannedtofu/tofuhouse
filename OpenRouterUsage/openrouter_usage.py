# -*- coding: utf-8 -*-
"""Weekly OpenRouter token-usage tracker — fetches the public rankings
chart API, accumulates history in Excel, pushes two panels to the
news_agent dashboard. No browser automation needed (unlike GAODE.py) —
the chart is backed by a plain public JSON endpoint.
"""

import os
import datetime
import traceback

import requests
import openpyxl

os.chdir(os.path.dirname(os.path.abspath(__file__)))

_LOG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "openrouter_run.log")


def _log(msg):
    with open(_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(f"[{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {msg}\n")


def parse_chart_data(payload: dict) -> list[dict]:
    """Parse the OpenRouter rankings API response into per-week records.

    Returns a list of {"date": "YYYY-MM-DD", "models": {model: tokens, ...}, "total": int}
    sorted by date ascending. "total" includes the "Others" bucket.
    """
    entries = payload["data"]["data"]
    records = []
    for entry in entries:
        models = dict(entry["ys"])
        records.append({
            "date": entry["x"],
            "models": models,
            "total": sum(models.values()),
        })
    records.sort(key=lambda r: r["date"])
    return records


def compute_deltas(records: list[dict]) -> list[dict]:
    """Given records sorted by date ascending (each with 'date' and 'total'),
    return rows annotated with wow_delta and wow_delta_of_delta.

    First row: both None (no prior week to diff against).
    Second row: wow_delta set, wow_delta_of_delta still None (no prior delta).
    """
    rows = []
    prev_total = None
    prev_delta = None
    for r in records:
        total = r["total"]
        delta = None if prev_total is None else total - prev_total
        delta_of_delta = None if (delta is None or prev_delta is None) else delta - prev_delta
        rows.append({
            "date": r["date"],
            "total": total,
            "wow_delta": delta,
            "wow_delta_of_delta": delta_of_delta,
        })
        prev_total = total
        prev_delta = delta
    return rows


def build_panels(rows: list[dict]) -> list[dict]:
    """Build the two dashboard panels from delta-annotated weekly rows.

    Panel 1: total token usage (single line) — the literal "total instead
    of OpenRouter's per-model stacked chart" the dashboard shows.
    Panel 2: week-over-week change — two lines, span_gaps=False so the
    chart doesn't draw a misleading line across the first 1-2 weeks
    where there's no prior data to diff against.
    """
    total_panel = {
        "type": "line",
        "title": "OpenRouter 周度 Token 总用量",
        "x_type": "date",
        "span_gaps": True,
        "datasets": [
            {
                "label": "Total Tokens",
                "data": [{"x": r["date"], "y": r["total"]} for r in rows],
            }
        ],
    }
    delta_panel = {
        "type": "line",
        "title": "Token 用量周环比变化",
        "x_type": "date",
        "span_gaps": False,
        "datasets": [
            {
                "label": "周环比变化 (WoW Δ)",
                "data": [{"x": r["date"], "y": r["wow_delta"]} for r in rows],
            },
            {
                "label": "环比变化的环比变化 (WoW Δ-of-Δ)",
                "data": [{"x": r["date"], "y": r["wow_delta_of_delta"]} for r in rows],
            },
        ],
    }
    return [total_panel, delta_panel]


_WEEKLY_TOTAL_HEADERS = ["Date", "Total Tokens", "WoW Δ", "WoW Δ-of-Δ"]


def upsert_weekly_total_sheet(wb, rows: list[dict]) -> None:
    """Upsert rows into the 'Weekly Total' sheet, keyed by Date.

    Dates not present in `rows` (e.g. weeks that rolled out of the API's
    rolling window) are preserved untouched — this is what gives the
    Excel file permanent history beyond what OpenRouter currently exposes.

    Uses merge-then-rebuild (dict by date, then clear + rewrite sorted)
    rather than in-place row updates/inserts, since keeping a growing
    sheet sorted and free of duplicate dates via targeted insert/update
    is more error-prone than rebuilding from a dict each time.
    """
    if "Weekly Total" not in wb.sheetnames:
        ws = wb.create_sheet("Weekly Total")
    else:
        ws = wb["Weekly Total"]

    existing: dict[str, list] = {}
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 5)]
        if vals[0]:
            existing[str(vals[0])] = vals

    for r in rows:
        existing[r["date"]] = [r["date"], r["total"], r["wow_delta"], r["wow_delta_of_delta"]]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_WEEKLY_TOTAL_HEADERS)
    for date_str in sorted(existing.keys()):
        ws.append(existing[date_str])


_MODEL_DETAIL_HEADERS = ["Date", "Model", "Tokens"]


def upsert_model_detail_sheet(wb, records: list[dict]) -> None:
    """Upsert per-model rows into the 'Per-Model Detail' sheet, keyed by Date.

    All existing rows for a date present in `records` are replaced with the
    fresh set (handles models appearing/disappearing within a week). Dates
    not present in `records` are left untouched.

    Uses merge-then-rebuild, but at the per-date *block* level rather than
    per-row like 'Weekly Total': a date here spans multiple rows (one per
    model), and the set of models for a given week can grow or shrink between
    runs, so matching/updating individual rows would leave stale model rows
    behind. Replacing the whole block for a date sidesteps that.
    """
    if "Per-Model Detail" not in wb.sheetnames:
        ws = wb.create_sheet("Per-Model Detail")
    else:
        ws = wb["Per-Model Detail"]

    incoming_dates = {r["date"] for r in records}

    kept_rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 4)]
        if vals[0] and str(vals[0]) not in incoming_dates:
            kept_rows.append(vals)

    new_rows = [
        [r["date"], model, tokens]
        for r in records
        for model, tokens in r["models"].items()
    ]
    # Sorted by date only — model order within a date block follows
    # insertion order (kept rows first, then this run's records), not a
    # secondary sort key. That's fine; model order within a week is cosmetic.
    all_rows = sorted(kept_rows + new_rows, key=lambda row: row[0])

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_MODEL_DETAIL_HEADERS)
    for row in all_rows:
        ws.append(row)


def _accumulated_totals(wb, records: list[dict]) -> list[dict]:
    """Merge this run's per-week totals with totals already stored in the
    Weekly Total sheet, so delta recompute covers full accumulated history
    (not just the weeks returned by this run's API call).
    """
    totals: dict[str, int] = {}
    if "Weekly Total" in wb.sheetnames:
        ws = wb["Weekly Total"]
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            total_val = ws.cell(row=row_idx, column=2).value
            if date_val is not None and total_val is not None:
                totals[str(date_val)] = total_val
    for r in records:
        totals[r["date"]] = r["total"]
    return [{"date": d, "total": totals[d]} for d in sorted(totals.keys())]


_API_BASE = "http://47.239.66.248"
_API_KEY = "b1445fd803c77c5bff4b0eeced29f5b84c752d0bbd6642f89bd44c732a1646fa"
_SCRIPT_NAME = "openrouter_usage"


def fetch_chart_data() -> dict:
    resp = requests.get(
        "https://openrouter.ai/api/frontend/v1/rankings/model-rankings-chart",
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def push_to_dashboard(rows: list[dict], excel_path: str) -> None:
    panels = build_panels(rows)
    sess = requests.Session()
    sess.trust_env = False  # don't pick up Windows proxy settings (see GAODE.py)
    sess.post(
        f"{_API_BASE}/api/report",
        headers={"X-API-Key": _API_KEY},
        json={
            "script": _SCRIPT_NAME,
            "status": "ok",
            "expected_interval_hours": 168,  # 7 days
            "panels": panels,
        },
        timeout=30,
    ).raise_for_status()
    with open(excel_path, "rb") as f:
        sess.post(
            f"{_API_BASE}/api/report/{_SCRIPT_NAME}/excel",
            headers={"X-API-Key": _API_KEY},
            files={"file": ("openrouter_usage.xlsx", f)},
            timeout=60,
        ).raise_for_status()


def main():
    _log("=" * 60)
    _log("Script started")

    try:
        payload = fetch_chart_data()
    except Exception:
        _log("Fetch FAILED:\n" + traceback.format_exc())
        raise

    records = parse_chart_data(payload)
    _log(f"Fetched {len(records)} weeks from API")

    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "openrouter_usage.xlsx")
    wb = openpyxl.load_workbook(excel_path) if os.path.exists(excel_path) else openpyxl.Workbook()

    upsert_model_detail_sheet(wb, records)

    full_history = _accumulated_totals(wb, records)
    rows = compute_deltas(full_history)
    upsert_weekly_total_sheet(wb, rows)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]  # drop openpyxl's default blank sheet

    wb.save(excel_path)
    _log(f"Excel saved: {excel_path}")

    try:
        push_to_dashboard(rows, excel_path)
        _log("Dashboard push OK")
    except Exception as e:
        _log(f"Dashboard push FAILED:\n{traceback.format_exc()}")
        print(f"Dashboard push failed (non-fatal): {e}")


if __name__ == "__main__":
    main()

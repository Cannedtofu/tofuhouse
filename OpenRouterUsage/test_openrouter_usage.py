import openpyxl

from openrouter_usage import parse_chart_data, compute_deltas, build_panels, upsert_weekly_total_sheet
from openrouter_usage import upsert_model_detail_sheet
from openrouter_usage import _accumulated_totals


def _read_weekly_total_rows(wb):
    ws = wb["Weekly Total"]
    return [
        [ws.cell(row=r, column=c).value for c in range(1, 5)]
        for r in range(2, ws.max_row + 1)
    ]


def test_parse_chart_data_computes_total_and_sorts():
    payload = {
        "data": {
            "data": [
                {"x": "2025-07-07", "ys": {"model-a": 100, "Others": 50}},
                {"x": "2025-06-30", "ys": {"model-a": 80, "model-b": 20, "Others": 10}},
            ]
        }
    }
    records = parse_chart_data(payload)
    assert [r["date"] for r in records] == ["2025-06-30", "2025-07-07"]
    assert records[0]["total"] == 110
    assert records[1]["total"] == 150
    assert records[0]["models"] == {"model-a": 80, "model-b": 20, "Others": 10}


def test_compute_deltas_first_row_has_no_delta():
    rows = compute_deltas([{"date": "2025-06-30", "total": 100}])
    assert rows[0]["wow_delta"] is None
    assert rows[0]["wow_delta_of_delta"] is None


def test_compute_deltas_second_row_has_delta_but_no_delta_of_delta():
    rows = compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ])
    assert rows[1]["wow_delta"] == 50
    assert rows[1]["wow_delta_of_delta"] is None


def test_compute_deltas_third_row_has_both():
    rows = compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
        {"date": "2025-07-14", "total": 130},
    ])
    assert rows[2]["wow_delta"] == -20
    assert rows[2]["wow_delta_of_delta"] == -20 - 50  # -70


def test_build_panels_total_panel_structure():
    rows = compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ])
    panels = build_panels(rows)
    total_panel = panels[0]
    assert total_panel["type"] == "line"
    assert total_panel["x_type"] == "date"
    assert total_panel["span_gaps"] is True
    assert total_panel["datasets"] == [
        {"label": "Total Tokens", "data": [
            {"x": "2025-06-30", "y": 100},
            {"x": "2025-07-07", "y": 150},
        ]}
    ]


def test_build_panels_delta_panel_has_two_datasets_with_nulls():
    rows = compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ])
    panels = build_panels(rows)
    delta_panel = panels[1]
    assert delta_panel["type"] == "line"
    assert delta_panel["x_type"] == "date"
    assert delta_panel["span_gaps"] is False
    assert len(delta_panel["datasets"]) == 2
    wow_delta_data = delta_panel["datasets"][0]["data"]
    assert wow_delta_data[0]["y"] is None   # first week: no prior week
    assert wow_delta_data[1]["y"] == 50
    wow_delta_of_delta_data = delta_panel["datasets"][1]["data"]
    assert wow_delta_of_delta_data[0]["y"] is None
    assert wow_delta_of_delta_data[1]["y"] is None  # second week: no prior delta yet


def test_upsert_weekly_total_sheet_inserts_new_rows():
    wb = openpyxl.Workbook()
    rows = compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ])
    upsert_weekly_total_sheet(wb, rows)
    data = _read_weekly_total_rows(wb)
    assert data == [
        ["2025-06-30", 100, None, None],
        ["2025-07-07", 150, 50, None],
    ]


def test_upsert_weekly_total_sheet_overwrites_existing_date_and_preserves_others():
    wb = openpyxl.Workbook()
    upsert_weekly_total_sheet(wb, compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ]))
    # Second run: API revises 2025-07-07's total, adds a new week.
    # 2025-06-30 is no longer in the API's window but must be preserved.
    upsert_weekly_total_sheet(wb, compute_deltas([
        {"date": "2025-07-07", "total": 160},
        {"date": "2025-07-14", "total": 200},
    ]))
    data = _read_weekly_total_rows(wb)
    assert data == [
        ["2025-06-30", 100, None, None],   # preserved, untouched
        ["2025-07-07", 160, None, None],   # overwritten with revised total
        ["2025-07-14", 200, 40, None],     # newly added
    ]


def _read_model_detail_rows(wb):
    ws = wb["Per-Model Detail"]
    return [
        [ws.cell(row=r, column=c).value for c in range(1, 4)]
        for r in range(2, ws.max_row + 1)
    ]


def test_upsert_model_detail_sheet_inserts_and_refreshes_by_date():
    wb = openpyxl.Workbook()
    upsert_model_detail_sheet(wb, [
        {"date": "2025-06-30", "models": {"model-a": 80, "model-b": 20}, "total": 100},
    ])
    assert _read_model_detail_rows(wb) == [
        ["2025-06-30", "model-a", 80],
        ["2025-06-30", "model-b", 20],
    ]

    # Second run: 2025-06-30's model breakdown is refreshed (model-b dropped,
    # model-c appears); a new week is added. Both should be reflected exactly,
    # nothing duplicated.
    upsert_model_detail_sheet(wb, [
        {"date": "2025-06-30", "models": {"model-a": 90, "model-c": 5}, "total": 95},
        {"date": "2025-07-07", "models": {"model-a": 100}, "total": 100},
    ])
    assert _read_model_detail_rows(wb) == [
        ["2025-06-30", "model-a", 90],
        ["2025-06-30", "model-c", 5],
        ["2025-07-07", "model-a", 100],
    ]


def test_accumulated_totals_merges_existing_and_new():
    wb = openpyxl.Workbook()
    upsert_weekly_total_sheet(wb, compute_deltas([
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
    ]))
    # This run's fetch only returned the newest 2 weeks (older one rolled
    # out of the API's window) — but 2025-06-30 must still be included.
    new_records = [
        {"date": "2025-07-07", "total": 150},
        {"date": "2025-07-14", "total": 200},
    ]
    merged = _accumulated_totals(wb, new_records)
    assert merged == [
        {"date": "2025-06-30", "total": 100},
        {"date": "2025-07-07", "total": 150},
        {"date": "2025-07-14", "total": 200},
    ]

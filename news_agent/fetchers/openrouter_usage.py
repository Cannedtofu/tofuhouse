"""Fetches weekly OpenRouter token-usage data and builds dashboard panels.

Public API:
  run_openrouter_usage_fetch() -> (panels: list[dict], excel_bytes: bytes)

The OpenRouter rankings chart (https://openrouter.ai/rankings) is backed by a
fully public JSON API — no auth, no browser automation needed. The API
returns a rolling window of weekly history (currently ~52 weeks); this module
accumulates that into a permanent record stored as an Excel blob in the
script_files table (via db.upsert_script_file), since the API's window can
silently drop old weeks over time.
"""

import io
import logging

import requests
import openpyxl

import db

log = logging.getLogger(__name__)

_OPENROUTER_API_URL = "https://openrouter.ai/api/frontend/v1/rankings/model-rankings-chart"
_SCRIPT_NAME = "openrouter_usage"


def fetch_chart_data() -> dict:
    resp = requests.get(
        _OPENROUTER_API_URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def parse_chart_data(payload: dict) -> list[dict]:
    """Parse the OpenRouter rankings API response into per-week records.

    Returns a list of {"date": "YYYY-MM-DD", "models": {model: tokens, ...}, "total": int}
    sorted by date ascending. "total" includes the "Others" bucket.
    """
    entries = payload["data"]["data"]
    records = []
    for entry in entries:
        models = dict(entry["ys"])
        records.append({
            "date": entry["x"],
            "models": models,
            "total": sum(models.values()),
        })
    records.sort(key=lambda r: r["date"])
    return records


def _trailing_avg(values: list, window: int) -> list:
    """Return a list the same length as `values`, where index i holds the
    average of values[i-window+1 : i+1], or None if fewer than `window`
    values exist yet (i.e. for the first window-1 entries)."""
    n = len(values)
    out = [None] * n
    for i in range(n):
        if i >= window - 1:
            out[i] = sum(values[i - window + 1:i + 1]) / window
    return out


def compute_deltas(records: list[dict]) -> list[dict]:
    """Given records sorted by date ascending (each with 'date' and 'total'),
    return rows annotated with several derived metrics:

    - total_wow_pct_change: week-over-week growth rate of the raw total —
      (this week's total / last week's total) - 1. None for the first row
      (no prior week) or if the prior total is exactly zero.
    - wow_delta_avg4: week-over-week change of the 4-week trailing average
      of total tokens (avg4[i] - avg4[i-1]) — smooths out single-week noise
      compared to diffing the raw weekly totals directly. None until there
      are at least 5 weeks of history.
    - avg3: 3-week trailing average of total tokens. None until there are
      at least 3 weeks of history.
    - avg3_wow_pct_change: week-over-week growth rate of avg3 — smooths out
      single-week noise compared to diffing the raw totals' pct change
      directly. None until there are at least 4 weeks of history (avg3
      itself needs 3 trailing weeks, and this diffs two consecutive avg3
      values), or if the prior avg3 is exactly zero.
    """
    totals = [r["total"] for r in records]
    n = len(totals)

    avg3 = _trailing_avg(totals, 3)
    avg4 = _trailing_avg(totals, 4)

    rows = []
    for i, r in enumerate(records):
        total = r["total"]

        pct_change = None
        if i >= 1 and totals[i - 1] != 0:
            pct_change = (total / totals[i - 1]) - 1

        wow_delta_avg4 = None
        if avg4[i] is not None and i >= 1 and avg4[i - 1] is not None:
            wow_delta_avg4 = avg4[i] - avg4[i - 1]

        avg3_wow_pct_change = None
        if (avg3[i] is not None and i >= 1 and avg3[i - 1] is not None
                and avg3[i - 1] != 0):
            avg3_wow_pct_change = (avg3[i] / avg3[i - 1]) - 1

        rows.append({
            "date": r["date"],
            "total": total,
            "avg3": avg3[i],
            "wow_delta_avg4": wow_delta_avg4,
            "total_wow_pct_change": pct_change,
            "avg3_wow_pct_change": avg3_wow_pct_change,
        })
    return rows


def build_panels(rows: list[dict]) -> list[dict]:
    """Build the two dashboard panels from delta-annotated weekly rows.

    Panel 1: total token usage (single line) — the literal "total instead
    of OpenRouter's per-model stacked chart" the dashboard shows.
    Panel 2: week-over-week change — two lines, span_gaps=False so the
    chart doesn't draw a misleading line across the first 1-2 weeks
    where there's no prior data to diff against.
    """
    total_panel = {
        "type": "line",
        "title": "OpenRouter 周度 Token 总用量",
        "x_type": "date",
        "span_gaps": True,
        "datasets": [
            {
                "label": "Total Tokens",
                "data": [{"x": r["date"], "y": r["total"]} for r in rows],
            },
            {
                "label": "3周均值 (3-Week Avg)",
                "axis": "left",
                "data": [{"x": r["date"], "y": r["avg3"]} for r in rows],
            },
            {
                # Independent right axis — a percentage shares no useful scale
                # with absolute token counts.
                "label": "3周均值环比变化率 (3-Week Avg WoW % Change)",
                "axis": "right",
                "format": "percent",
                "data": [{"x": r["date"], "y": r["avg3_wow_pct_change"]} for r in rows],
            },
        ],
    }
    delta_panel = {
        "type": "line",
        "title": "Token 用量周环比变化",
        "x_type": "date",
        "span_gaps": False,
        "datasets": [
            {
                "label": "4周均值环比变化 (4-Week Avg WoW Δ)",
                "axis": "left",
                "data": [{"x": r["date"], "y": r["wow_delta_avg4"]} for r in rows],
            },
            {
                # On its own (right) axis as a percentage — absolute token
                # counts and a ratio live on wildly different scales, so
                # sharing one axis would flatten one of the two lines.
                "label": "总量环比变化率 (Total Token WoW % Change)",
                "axis": "right",
                "format": "percent",
                "data": [{"x": r["date"], "y": r["total_wow_pct_change"]} for r in rows],
            },
        ],
    }
    return [total_panel, delta_panel]


_WEEKLY_TOTAL_HEADERS = [
    "Date", "Total Tokens", "4-Week Avg WoW Δ", "Total WoW % Change",
    "3-Week Avg", "3-Week Avg WoW % Change",
]


def upsert_weekly_total_sheet(wb, rows: list[dict]) -> None:
    """Upsert rows into the 'Weekly Total' sheet, keyed by Date.

    Dates not present in `rows` (e.g. weeks that rolled out of the API's
    rolling window) are preserved untouched — this is what gives the
    Excel file permanent history beyond what OpenRouter currently exposes.

    Uses merge-then-rebuild (dict by date, then clear + rewrite sorted)
    rather than in-place row updates/inserts, since keeping a growing
    sheet sorted and free of duplicate dates via targeted insert/update
    is more error-prone than rebuilding from a dict each time.

    Every date in `rows` always overwrites the stored row regardless of
    whether the value changed — but a mismatch against the *raw* Total
    Tokens figure (not the derived WoW columns, which legitimately cascade
    whenever any upstream week is corrected) is logged, since that's the
    signal that OpenRouter itself revised a historical number.
    """
    if "Weekly Total" not in wb.sheetnames:
        ws = wb.create_sheet("Weekly Total")
    else:
        ws = wb["Weekly Total"]

    existing: dict[str, list] = {}
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 7)]
        if vals[0]:
            existing[str(vals[0])] = vals

    for r in rows:
        prior = existing.get(r["date"])
        if prior is not None and prior[1] != r["total"]:
            log.info(
                "OpenRouter usage: %s total revised by source — %s -> %s",
                r["date"], prior[1], r["total"],
            )
        existing[r["date"]] = [
            r["date"], r["total"], r["wow_delta_avg4"], r["total_wow_pct_change"],
            r["avg3"], r["avg3_wow_pct_change"],
        ]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_WEEKLY_TOTAL_HEADERS)
    for date_str in sorted(existing.keys()):
        ws.append(existing[date_str])


_MODEL_DETAIL_HEADERS = ["Date", "Model", "Tokens"]


def upsert_model_detail_sheet(wb, records: list[dict]) -> None:
    """Upsert per-model rows into the 'Per-Model Detail' sheet, keyed by Date.

    All existing rows for a date present in `records` are replaced with the
    fresh set (handles models appearing/disappearing within a week). Dates
    not present in `records` are left untouched.

    Uses merge-then-rebuild, but at the per-date *block* level rather than
    per-row like 'Weekly Total': a date here spans multiple rows (one per
    model), and the set of models for a given week can grow or shrink between
    runs, so matching/updating individual rows would leave stale model rows
    behind. Replacing the whole block for a date sidesteps that.

    Every (date, model) pair in `records` always overwrites the stored row
    regardless of whether the value changed — but a token-count mismatch
    against what was already stored is logged, since that's the signal
    that OpenRouter revised a historical number rather than this just
    being the model's first appearance for that week.
    """
    if "Per-Model Detail" not in wb.sheetnames:
        ws = wb.create_sheet("Per-Model Detail")
    else:
        ws = wb["Per-Model Detail"]

    incoming_dates = {r["date"] for r in records}

    existing_tokens: dict[tuple, int] = {}
    kept_rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = [ws.cell(row=row_idx, column=c).value for c in range(1, 4)]
        if not vals[0]:
            continue
        existing_tokens[(str(vals[0]), vals[1])] = vals[2]
        if str(vals[0]) not in incoming_dates:
            kept_rows.append(vals)

    for r in records:
        for model, tokens in r["models"].items():
            prior_tokens = existing_tokens.get((r["date"], model))
            if prior_tokens is not None and prior_tokens != tokens:
                log.info(
                    "OpenRouter usage: %s / %s tokens revised by source — %s -> %s",
                    r["date"], model, prior_tokens, tokens,
                )

    new_rows = [
        [r["date"], model, tokens]
        for r in records
        for model, tokens in r["models"].items()
    ]
    # Sorted by date only — model order within a date block follows
    # insertion order (kept rows first, then this run's records), not a
    # secondary sort key. That's fine; model order within a week is cosmetic.
    all_rows = sorted(kept_rows + new_rows, key=lambda row: row[0])

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_MODEL_DETAIL_HEADERS)
    for row in all_rows:
        ws.append(row)


def _accumulated_totals(wb, records: list[dict]) -> list[dict]:
    """Merge this run's per-week totals with totals already stored in the
    Weekly Total sheet, so delta recompute covers full accumulated history
    (not just the weeks returned by this run's API call).
    """
    totals: dict[str, int] = {}
    if "Weekly Total" in wb.sheetnames:
        ws = wb["Weekly Total"]
        for row_idx in range(2, ws.max_row + 1):
            date_val = ws.cell(row=row_idx, column=1).value
            total_val = ws.cell(row=row_idx, column=2).value
            if date_val is not None and total_val is not None:
                totals[str(date_val)] = total_val
    for r in records:
        totals[r["date"]] = r["total"]
    return [{"date": d, "total": totals[d]} for d in sorted(totals.keys())]


def run_openrouter_usage_fetch() -> tuple[list[dict], bytes]:
    """Fetch live data, merge into the accumulated Excel history (stored as
    a DB blob, not a local file — this runs in-process on the server), and
    return (panels, excel_bytes) for the caller to push into script_reports
    / script_files directly (no HTTP round-trip needed since we're already
    in the same process as the dashboard).
    """
    payload = fetch_chart_data()
    records = parse_chart_data(payload)
    log.info("OpenRouter usage: fetched %d weeks from API", len(records))

    existing_file = db.get_script_file(_SCRIPT_NAME)
    if existing_file:
        wb = openpyxl.load_workbook(io.BytesIO(existing_file["file_data"]))
    else:
        wb = openpyxl.Workbook()

    upsert_model_detail_sheet(wb, records)

    full_history = _accumulated_totals(wb, records)
    rows = compute_deltas(full_history)
    upsert_weekly_total_sheet(wb, rows)

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]  # drop openpyxl's default blank sheet

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    panels = build_panels(rows)
    return panels, excel_bytes

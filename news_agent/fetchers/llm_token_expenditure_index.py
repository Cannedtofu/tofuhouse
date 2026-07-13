from __future__ import annotations

import io
import json
import logging
import re
from datetime import datetime, timezone

import openpyxl
import requests

import db

log = logging.getLogger(__name__)

_SILICON_EMBED_URL = "https://portal.silicondata.com/token-index-chart"
_TRAKTOKEN_CANONICAL_URL = "https://www.traktoken.com/spend-index"
_TRAKTOKEN_HISTORY_URL = "https://www.traktoken.com/api/index/history"
_SCRIPT_NAME = "LLM Token Expenditure Index"
_SILICON_SHEET = "Silicon Data"
_TRAKTOKEN_SHEET = "TrakToken"
_LEGACY_SHEET = "Daily Index"
_SILICON_HEADERS = ["As Of Date", "USD / 1M Tokens", "Fetched At UTC", "Source URL"]
_TRAKTOKEN_HEADERS = [
    "As Of Date",
    "USD / 1M Tokens (MA7)",
    "USD / 1M Tokens (Raw)",
    "TTSI (MA7)",
    "TTSI (Raw)",
    "Fetched At UTC",
    "Source URL",
]


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def fetch_silicon_snapshot_series() -> list[dict]:
    resp = requests.get(
        _SILICON_EMBED_URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=30,
    )
    resp.raise_for_status()
    return parse_silicon_snapshot_series(resp.text)


def parse_silicon_snapshot_series(html: str) -> list[dict]:
    series_match = re.search(r'\\"indexes\\":\{([^{}]+)\}', html)
    if not series_match:
        raise ValueError("Could not parse Silicon Data token index series")

    fetched_at = _utc_now_iso()
    series_json = "{" + series_match.group(1).replace('\\"', '"') + "}"
    indexes = json.loads(series_json)
    rows = []
    for as_of_date, value in sorted(indexes.items()):
        rows.append({
            "as_of_date": as_of_date,
            "usd_per_m_tokens": float(value),
            "fetched_at": fetched_at,
            "source_url": _SILICON_EMBED_URL,
        })
    return rows


def fetch_traktoken_history_series(date_from: str = "2025-12-01", date_to: str | None = None) -> list[dict]:
    if not date_to:
        date_to = datetime.now(timezone.utc).date().isoformat()

    resp = requests.get(
        _TRAKTOKEN_HISTORY_URL,
        params={"from": date_from, "to": date_to},
        headers={"User-Agent": "Mozilla/5.0", "Accept": "application/json"},
        timeout=30,
    )
    resp.raise_for_status()
    payload = resp.json()
    if not payload.get("success"):
        raise ValueError(f"TrakToken history API returned unsuccessful payload: {payload}")

    fetched_at = _utc_now_iso()
    rows = []
    for item in payload.get("data") or []:
        as_of_date = item.get("date")
        ma7_price = item.get("spend_price_usd_ma7")
        raw_price = item.get("spend_price_usd")
        ttsi_ma7 = item.get("ttsi_ma7")
        ttsi_raw = item.get("ttsi")
        if not as_of_date or ma7_price is None:
            continue
        rows.append({
            "as_of_date": str(as_of_date),
            "usd_per_m_tokens": float(ma7_price),
            "usd_per_m_tokens_raw": float(raw_price) if raw_price is not None else None,
            "ttsi": float(ttsi_raw) if ttsi_raw is not None else None,
            "ttsi_ma7": float(ttsi_ma7) if ttsi_ma7 is not None else None,
            "fetched_at": fetched_at,
            "source_url": _TRAKTOKEN_CANONICAL_URL,
        })
    return sorted(rows, key=lambda row: row["as_of_date"])


def _load_legacy_silicon_rows(wb) -> dict[str, dict]:
    if _LEGACY_SHEET not in wb.sheetnames:
        return {}

    ws = wb[_LEGACY_SHEET]
    existing: dict[str, dict] = {}
    for row_idx in range(2, ws.max_row + 1):
        as_of_date = ws.cell(row=row_idx, column=1).value
        index_value = ws.cell(row=row_idx, column=2).value
        fetched_at = ws.cell(row=row_idx, column=3).value
        source_url = ws.cell(row=row_idx, column=4).value
        if not as_of_date or index_value is None:
            continue
        existing[str(as_of_date)] = {
            "as_of_date": str(as_of_date),
            "usd_per_m_tokens": float(index_value),
            "fetched_at": str(fetched_at or ""),
            "source_url": str(source_url or _SILICON_EMBED_URL),
        }
    return existing


def _load_silicon_rows(wb) -> dict[str, dict]:
    if _SILICON_SHEET not in wb.sheetnames:
        return _load_legacy_silicon_rows(wb)

    ws = wb[_SILICON_SHEET]
    existing: dict[str, dict] = {}
    for row_idx in range(2, ws.max_row + 1):
        as_of_date = ws.cell(row=row_idx, column=1).value
        usd_per_m_tokens = ws.cell(row=row_idx, column=2).value
        fetched_at = ws.cell(row=row_idx, column=3).value
        source_url = ws.cell(row=row_idx, column=4).value
        if not as_of_date or usd_per_m_tokens is None:
            continue
        existing[str(as_of_date)] = {
            "as_of_date": str(as_of_date),
            "usd_per_m_tokens": float(usd_per_m_tokens),
            "fetched_at": str(fetched_at or ""),
            "source_url": str(source_url or _SILICON_EMBED_URL),
        }
    return existing


def _load_traktoken_rows(wb) -> dict[str, dict]:
    if _TRAKTOKEN_SHEET not in wb.sheetnames:
        return {}

    ws = wb[_TRAKTOKEN_SHEET]
    existing: dict[str, dict] = {}
    for row_idx in range(2, ws.max_row + 1):
        as_of_date = ws.cell(row=row_idx, column=1).value
        usd_ma7 = ws.cell(row=row_idx, column=2).value
        usd_raw = ws.cell(row=row_idx, column=3).value
        ttsi_ma7 = ws.cell(row=row_idx, column=4).value
        ttsi_raw = ws.cell(row=row_idx, column=5).value
        fetched_at = ws.cell(row=row_idx, column=6).value
        source_url = ws.cell(row=row_idx, column=7).value
        if not as_of_date or usd_ma7 is None:
            continue
        existing[str(as_of_date)] = {
            "as_of_date": str(as_of_date),
            "usd_per_m_tokens": float(usd_ma7),
            "usd_per_m_tokens_raw": float(usd_raw) if usd_raw is not None else None,
            "ttsi_ma7": float(ttsi_ma7) if ttsi_ma7 is not None else None,
            "ttsi": float(ttsi_raw) if ttsi_raw is not None else None,
            "fetched_at": str(fetched_at or ""),
            "source_url": str(source_url or _TRAKTOKEN_CANONICAL_URL),
        }
    return existing


def _merge_rows(existing_rows: dict[str, dict], snapshots: list[dict], source_label: str, value_key: str = "usd_per_m_tokens") -> list[dict]:
    merged = dict(existing_rows)
    for snapshot in snapshots:
        prior = merged.get(snapshot["as_of_date"])
        if prior is not None and prior.get(value_key) != snapshot.get(value_key):
            log.info(
                "%s token expenditure index: %s revised by source - %s -> %s",
                source_label,
                snapshot["as_of_date"],
                prior.get(value_key),
                snapshot.get(value_key),
            )
        merged[snapshot["as_of_date"]] = snapshot
    return [merged[key] for key in sorted(merged.keys())]


def _upsert_silicon_sheet(wb, rows: list[dict]) -> None:
    if _SILICON_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_SILICON_SHEET)
    else:
        ws = wb[_SILICON_SHEET]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_SILICON_HEADERS)
    for row in rows:
        ws.append([
            row["as_of_date"],
            row["usd_per_m_tokens"],
            row["fetched_at"],
            row["source_url"],
        ])


def _upsert_traktoken_sheet(wb, rows: list[dict]) -> None:
    if _TRAKTOKEN_SHEET not in wb.sheetnames:
        ws = wb.create_sheet(_TRAKTOKEN_SHEET)
    else:
        ws = wb[_TRAKTOKEN_SHEET]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_TRAKTOKEN_HEADERS)
    for row in rows:
        ws.append([
            row["as_of_date"],
            row["usd_per_m_tokens"],
            row.get("usd_per_m_tokens_raw"),
            row.get("ttsi_ma7"),
            row.get("ttsi"),
            row["fetched_at"],
            row["source_url"],
        ])


def build_panels(silicon_rows: list[dict], traktoken_rows: list[dict]) -> list[dict]:
    datasets = []
    if silicon_rows:
        datasets.append({
            "label": "Silicon Data",
            "color": "#0d6efd",
            "data": [
                {"x": row["as_of_date"], "y": row["usd_per_m_tokens"]}
                for row in silicon_rows
            ],
        })
    if traktoken_rows:
        datasets.append({
            "label": "TrakToken (7D avg)",
            "color": "#fd7e14",
            "data": [
                {"x": row["as_of_date"], "y": row["usd_per_m_tokens"]}
                for row in traktoken_rows
            ],
        })

    if not datasets:
        raise ValueError("No LLM token expenditure index datasets available")

    return [{
        "type": "line",
        "title": "LLM Token Expenditure Index",
        "x_type": "date",
        "span_gaps": True,
        "datasets": datasets,
    }]


def run_llm_token_expenditure_index_fetch() -> tuple[list[dict], bytes]:
    existing_file = db.get_script_file(_SCRIPT_NAME)
    if existing_file:
        wb = openpyxl.load_workbook(io.BytesIO(existing_file["file_data"]))
    else:
        wb = openpyxl.Workbook()

    silicon_existing = _load_silicon_rows(wb)
    traktoken_existing = _load_traktoken_rows(wb)

    silicon_rows: list[dict]
    traktoken_rows: list[dict]

    silicon_error: Exception | None = None
    try:
        silicon_snapshots = fetch_silicon_snapshot_series()
        if not silicon_snapshots:
            raise ValueError("Silicon Data token index returned no rows")
        log.info(
            "LLM token expenditure index: Silicon Data fetched %d row(s) from %s to %s",
            len(silicon_snapshots),
            silicon_snapshots[0]["as_of_date"],
            silicon_snapshots[-1]["as_of_date"],
        )
        silicon_rows = _merge_rows(silicon_existing, silicon_snapshots, "Silicon Data")
    except Exception as exc:
        silicon_error = exc
        silicon_rows = [silicon_existing[key] for key in sorted(silicon_existing.keys())]
        log.warning("Silicon Data token index fetch failed, keeping cached history: %s", exc)

    traktoken_error: Exception | None = None
    try:
        traktoken_snapshots = fetch_traktoken_history_series()
        if not traktoken_snapshots:
            raise ValueError("TrakToken history API returned no rows")
        log.info(
            "LLM token expenditure index: TrakToken fetched %d row(s) from %s to %s",
            len(traktoken_snapshots),
            traktoken_snapshots[0]["as_of_date"],
            traktoken_snapshots[-1]["as_of_date"],
        )
        traktoken_rows = _merge_rows(traktoken_existing, traktoken_snapshots, "TrakToken")
    except Exception as exc:
        traktoken_error = exc
        traktoken_rows = [traktoken_existing[key] for key in sorted(traktoken_existing.keys())]
        log.warning("TrakToken token index fetch failed, keeping cached history: %s", exc)

    if not silicon_rows and silicon_error is not None:
        raise silicon_error
    if not traktoken_rows and traktoken_error is not None:
        raise traktoken_error

    _upsert_silicon_sheet(wb, silicon_rows)
    _upsert_traktoken_sheet(wb, traktoken_rows)

    if _LEGACY_SHEET in wb.sheetnames and _LEGACY_SHEET != _SILICON_SHEET:
        del wb[_LEGACY_SHEET]
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    panels = build_panels(silicon_rows, traktoken_rows)
    return panels, excel_bytes

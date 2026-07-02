"""Fetches Vercel AI Gateway Labs leaderboard data for the dashboard.

Public API:
  run_vercel_labs_fetch() -> (panels: list[dict], excel_bytes: bytes)

Uses Vercel's public leaderboard export endpoint rather than scraping page
markup. The `labs` dataset is updated daily, so this fetcher stores a growing
Excel history in the DB-backed script_files table and refreshes on a 24-hour
cadence.
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict

import openpyxl
import requests

import db

log = logging.getLogger(__name__)

_VERCEL_API_URL = "https://vercel.com/api/ai/leaderboard-export?dataset=labs"
_SCRIPT_NAME = "vercel_labs"
_RAW_HEADERS = ["Date", "Metric", "Lab", "Share Percent", "Modality"]
_TARGET_METRICS = ("tokens", "spend")


def fetch_chart_data() -> dict:
    resp = requests.get(
        _VERCEL_API_URL,
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()


def parse_chart_data(payload: dict) -> list[dict]:
    rows = []
    for row in payload.get("rows", []):
        if row.get("metric") not in _TARGET_METRICS:
            continue
        rows.append({
            "date": row["date"],
            "metric": row["metric"],
            "lab": row["name"],
            "share_percent": float(row["share_percent"]),
            "modality": row.get("modality", "all"),
        })
    rows.sort(key=lambda r: (r["date"], r["metric"], r["lab"]))
    return rows


def _load_existing_rows(wb) -> dict[tuple[str, str, str], dict]:
    if "Raw Rows" not in wb.sheetnames:
        return {}

    ws = wb["Raw Rows"]
    existing: dict[tuple[str, str, str], dict] = {}
    for row_idx in range(2, ws.max_row + 1):
        date_val = ws.cell(row=row_idx, column=1).value
        metric_val = ws.cell(row=row_idx, column=2).value
        lab_val = ws.cell(row=row_idx, column=3).value
        share_val = ws.cell(row=row_idx, column=4).value
        modality_val = ws.cell(row=row_idx, column=5).value
        if not date_val or not metric_val or not lab_val or share_val is None:
            continue
        existing[(str(date_val), str(metric_val), str(lab_val))] = {
            "date": str(date_val),
            "metric": str(metric_val),
            "lab": str(lab_val),
            "share_percent": float(share_val),
            "modality": str(modality_val or "all"),
        }
    return existing


def merge_rows(wb, incoming_rows: list[dict]) -> list[dict]:
    merged = _load_existing_rows(wb)
    for row in incoming_rows:
        key = (row["date"], row["metric"], row["lab"])
        prior = merged.get(key)
        if prior is not None and prior["share_percent"] != row["share_percent"]:
            log.info(
                "Vercel labs: %s / %s / %s revised by source - %s -> %s",
                row["date"], row["metric"], row["lab"],
                prior["share_percent"], row["share_percent"],
            )
        merged[key] = row
    return sorted(merged.values(), key=lambda r: (r["date"], r["metric"], r["lab"]))


def _records_by_metric(rows: list[dict]) -> dict[str, list[dict]]:
    grouped: dict[str, dict[str, dict[str, float]]] = defaultdict(lambda: defaultdict(dict))
    for row in rows:
        grouped[row["metric"]][row["date"]][row["lab"]] = row["share_percent"]

    out: dict[str, list[dict]] = {}
    for metric, by_date in grouped.items():
        out[metric] = [
            {"date": date_str, "labs": by_date[date_str]}
            for date_str in sorted(by_date.keys())
        ]
    return out


def _top_labs(records: list[dict], top_n: int = 8) -> list[str]:
    if not records:
        return []
    latest = records[-1]["labs"]
    return [
        name for name, _share in
        sorted(latest.items(), key=lambda item: item[1], reverse=True)[:top_n]
    ]


def _metric_title(metric: str) -> str:
    return "Token Volume" if metric == "tokens" else "Spend"


def build_panels(rows: list[dict]) -> list[dict]:
    by_metric = _records_by_metric(rows)
    panels = []

    for metric in _TARGET_METRICS:
        records = by_metric.get(metric, [])
        if not records:
            continue

        top_labs = _top_labs(records)
        datasets = []
        for lab in top_labs:
            datasets.append({
                "label": lab,
                "format": "percent",
                "data": [
                    {"x": record["date"], "y": record["labs"].get(lab, 0.0) / 100.0}
                    for record in records
                ],
            })

        if top_labs:
            datasets.append({
                "label": "others",
                "format": "percent",
                "data": [
                    {
                        "x": record["date"],
                        "y": sum(
                            share for lab, share in record["labs"].items()
                            if lab not in top_labs
                        ) / 100.0,
                    }
                    for record in records
                ],
            })

        panels.append({
            "type": "line",
            "title": f"Vercel Labs {_metric_title(metric)}",
            "x_type": "date",
            "span_gaps": True,
            "datasets": datasets,
        })

    return panels


def _upsert_raw_sheet(wb, rows: list[dict]) -> None:
    if "Raw Rows" not in wb.sheetnames:
        ws = wb.create_sheet("Raw Rows")
    else:
        ws = wb["Raw Rows"]

    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    ws.append(_RAW_HEADERS)
    for row in rows:
        ws.append([
            row["date"],
            row["metric"],
            row["lab"],
            row["share_percent"],
            row["modality"],
        ])


def _upsert_metric_sheet(wb, metric: str, records: list[dict]) -> None:
    title = "Token Share" if metric == "tokens" else "Spend Share"
    if title not in wb.sheetnames:
        ws = wb.create_sheet(title)
    else:
        ws = wb[title]

    labs = sorted({lab for record in records for lab in record["labs"].keys()})
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    ws.append(["Date"] + labs)
    for record in records:
        ws.append([record["date"]] + [record["labs"].get(lab) for lab in labs])


def run_vercel_labs_fetch() -> tuple[list[dict], bytes]:
    payload = fetch_chart_data()
    incoming_rows = parse_chart_data(payload)
    log.info("Vercel labs: fetched %d daily rows from API", len(incoming_rows))

    existing_file = db.get_script_file(_SCRIPT_NAME)
    if existing_file:
        wb = openpyxl.load_workbook(io.BytesIO(existing_file["file_data"]))
    else:
        wb = openpyxl.Workbook()

    merged_rows = merge_rows(wb, incoming_rows)
    _upsert_raw_sheet(wb, merged_rows)

    by_metric = _records_by_metric(merged_rows)
    for metric in _TARGET_METRICS:
        _upsert_metric_sheet(wb, metric, by_metric.get(metric, []))

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb["Sheet"]

    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()

    panels = build_panels(merged_rows)
    return panels, excel_bytes
